from openpyxl import load_workbook
from pathlib import Path
import json
wb_path = Path('../../asts/file-deps-methods.xlsx')
if not wb_path.exists():
    print('ERROR: workbook not found at', wb_path.resolve())
    raise SystemExit(1)
wb = load_workbook(wb_path, read_only=True, data_only=True)
# target file path as stored in sheet
target_end = 'PatientsController.cs'
result = {}
found = False
for name in wb.sheetnames:
    ws = wb[name]
    file_path = None
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if row[0] == 'FilePath':
            file_path = row[1]
            break
    if not file_path:
        continue
    if file_path.endswith(target_end):
        found = True
        rows = list(ws.iter_rows(values_only=True))
        methods_section = False
        current_method = None
        for r in rows:
            if r and r[0] == 'Methods and their call-chains:':
                methods_section = True
                continue
            if not methods_section:
                continue
            if r and r[0] and isinstance(r[0], str) and r[0].startswith('Method:'):
                current_method = r[0].split('Method: ',1)[1].strip()
                result[current_method] = {}
                continue
            if current_method and r and r[0] and str(r[0]).strip().startswith('  Level'):
                lvl = str(r[0]).strip()
                vals = r[1] or ''
                result[current_method][lvl] = [v.strip() for v in vals.split(';') if v.strip()]
            if current_method and r and all(cell is None for cell in r[:2]):
                current_method = None
        break
if not found:
    print('PatientsController sheet not found')
else:
    print(json.dumps(result, indent=2))
