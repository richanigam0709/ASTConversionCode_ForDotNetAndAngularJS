#!/usr/bin/env python3
"""
Generate an Excel report from source code files (interactive source root prompt if not provided).

Sheets:
 1. FileTypes - extension and total count (scans --source-root)
 2. Files - unique file_id and file path (declared namespaces and usings for C# files)
 3. Imports - mappings file_id,file_path -> imported_file_id,imported_file_path based on using->declared namespace matches

Usage (interactive):
  python tools/generate_imports_from_source.py
  (the script will prompt: "Enter source root path:")

Or with arguments:
  python tools/generate_imports_from_source.py --source-root "C:\path\to\repo" --output "asts_enhanced/file_imports_from_source.xlsx"

Requires: openpyxl
Install: pip install openpyxl
"""
import argparse
import re
import fnmatch
from pathlib import Path
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import sys
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, as_completed

NAMESPACE_RE = re.compile(r"^\s*namespace\s+([A-Za-z0-9_.]+)\s*(?:\{|;)")
USING_RE = re.compile(r"^\s*using\s+([A-Za-z0-9_.]+)\s*;")

# runtime options filled from CLI
GLOBAL_STRICT_USINGS = False

# when a file is only matched by a 'using' heuristic, skip it if its filename
# contains any of these keywords (helps filter out DependencyInjection/DbContext/Startup noise)
USING_IGNORE_KEYWORDS = ['dependencyinjection', 'databasecontext', 'startup', 'program']
# explicit user-provided filename exclusion patterns (case-insensitive substrings)
EXCLUDE_FILENAME_PATTERNS = []
# when True, skip matches that are only from a 'using' (include using matches only if
# the same file was also matched by another heuristic like param/new/di/method)
NO_USING_ONLY = False


def sanitize_sheet_name(name):
    return name[:31]


def strip_comments(text: str):
    """Remove // line comments and /* */ block comments to reduce false positives when regex-scanning."""
    if not text:
        return text
    # remove block comments
    text = re.sub(r"/\*.*?\*/", "", text, flags=re.S)
    # remove line comments
    text = re.sub(r"//.*?$", "", text, flags=re.M)
    return text


def autosize(ws, cols=None):
    if cols is None:
        cols = range(1, ws.max_column + 1)
    for col in cols:
        max_len = 0
        for cell in ws.iter_cols(min_col=col, max_col=col, values_only=False):
            for c in cell:
                try:
                    val = c.value or ''
                except Exception:
                    val = ''
                l = len(str(val))
                if l > max_len:
                    max_len = l
        width = min(max(10, max_len + 2), 120)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


def find_namespaces_and_usings(file_path: Path):
    nss = []
    us = []
    try:
        text = file_path.read_text(encoding='utf-8')
    except Exception:
        return nss, us
    for line in text.splitlines():
        m = NAMESPACE_RE.match(line)
        if m:
            nss.append(m.group(1))
        m2 = USING_RE.match(line)
        if m2:
            us.append(m2.group(1))
    # dedupe
    nss = list(dict.fromkeys(nss))
    us = list(dict.fromkeys(us))
    return nss, us


def find_declared_types_and_methods(file_path: Path):
    """Return tuple (classes, methods) where classes is list of declared class/type names and
    methods is list of declared method names in the file."""
    classes = []
    methods = []
    try:
        text = file_path.read_text(encoding='utf-8')
    except Exception:
        return classes, methods

    # class declarations
    for m in re.finditer(r"\bclass\s+([A-Za-z0-9_]+)", text):
        classes.append(m.group(1))

    # struct, record, interface, enum as types too
    for m in re.finditer(r"\b(struct|record|interface|enum)\s+([A-Za-z0-9_]+)", text):
        classes.append(m.group(2))

    # method declarations (simple heuristic)
    for m in re.finditer(r"\b(?:public|private|protected|internal|static|async|protected internal|internal protected)\s+[A-Za-z0-9_<>,\s\[\]]+\s+([A-Za-z0-9_]+)\s*\(", text):
        methods.append(m.group(1))

    # dedupe preserving order
    classes = list(dict.fromkeys(classes))
    methods = list(dict.fromkeys(methods))
    return classes, methods


def build_decl_indexes(source_files):
    """Build indexes: class_name -> file ids, method_name -> file ids"""
    class_idx = defaultdict(list)
    method_idx = defaultdict(list)
    for rec in source_files:
        fid = rec['id']
        path = Path(rec['path'])
        classes, methods = find_declared_types_and_methods(path)
        rec['declared_classes'] = classes
        rec['declared_methods'] = methods
        for c in classes:
            class_idx[c].append(fid)
    # dedupe
    nss = list(dict.fromkeys(nss))
    us = list(dict.fromkeys(us))
    return nss, us


def find_declared_types_and_methods(file_path: Path):
    """Return tuple (classes, methods) where classes is list of declared class/type names and
    methods is list of declared method names in the file."""
    classes = []
    methods = []
    try:
        text = file_path.read_text(encoding='utf-8')
    except Exception:
        return classes, methods

    # class declarations
    for m in re.finditer(r"\bclass\s+([A-Za-z0-9_]+)", text):
        classes.append(m.group(1))

    # struct, record, interface, enum as types too
    for m in re.finditer(r"\b(struct|record|interface|enum)\s+([A-Za-z0-9_]+)", text):
        classes.append(m.group(2))

    # method declarations (simple heuristic)
    for m in re.finditer(r"\b(?:public|private|protected|internal|static|async|protected internal|internal protected)\s+[A-Za-z0-9_<>,\s\[\]]+\s+([A-Za-z0-9_]+)\s*\(", text):
        methods.append(m.group(1))

    # dedupe preserving order
    classes = list(dict.fromkeys(classes))
    methods = list(dict.fromkeys(methods))
    return classes, methods


def build_decl_indexes(source_files):
    """Build indexes: class_name -> file ids, method_name -> file ids"""
    class_idx = defaultdict(list)
    method_idx = defaultdict(list)
    for rec in source_files:
        fid = rec['id']
        path = Path(rec['path'])
        classes, methods = find_declared_types_and_methods(path)
        rec['declared_classes'] = classes
        rec['declared_methods'] = methods
        for c in classes:
            class_idx[c].append(fid)
        for m in methods:
            method_idx[m].append(fid)
    return class_idx, method_idx


def build_decl_indexes_parallel(records, workers=None):
    """Parallel version of build_decl_indexes using ProcessPoolExecutor."""
    if workers is None:
        workers = max(1, multiprocessing.cpu_count() - 1)
    class_idx = defaultdict(list)
    method_idx = defaultdict(list)
    # submit parsing tasks
    with ProcessPoolExecutor(max_workers=workers) as ex:
        future_to_rec = {ex.submit(find_declared_types_and_methods, Path(rec['path'])): rec for rec in records}
        for fut in as_completed(future_to_rec):
            rec = future_to_rec[fut]
            try:
                classes, methods = fut.result()
            except Exception:
                classes, methods = [], []
            rec['declared_classes'] = classes
            rec['declared_methods'] = methods
            fid = rec['id']
            for c in classes:
                class_idx[c].append(fid)
            for m in methods:
                method_idx[m].append(fid)
    return class_idx, method_idx


def process_record_imports(args_tuple):
    """Worker function for import detection. args_tuple contains (rec, class_idx, method_idx, di_map, ns_to_ids)
    Returns (file_id, relpath, list_of_(imported_id, matched_by, matched_symbol))
    """
    rec, class_idx, method_idx, di_map, ns_to_ids = args_tuple
    matches = []

    # Use precomputed caches if available to avoid re-parsing the file
    text = rec.get('text', '')
    var_map = rec.get('var_map', {})
    param_field_types = rec.get('param_field_types', set())
    new_types = rec.get('new_types', set())
    invs = rec.get('invocations', [])

    seen = set()

    def should_skip_target(fid):
        """Return True if the target file's name matches any user-supplied exclude pattern."""
        try:
            records_list = globals().get('records', [])
            if not records_list:
                return False
            target = next((r for r in records_list if r['id'] == fid), None)
            if not target:
                return False
            fname = Path(target['path']).name.lower()
            for pat in EXCLUDE_FILENAME_PATTERNS:
                if pat and pat in fname:
                    return True
        except Exception:
            return False
        return False

    # 1) param/field/return types
    for t in param_field_types:
        if t in class_idx:
            for fid in class_idx[t]:
                if fid not in seen and fid != rec['id'] and not should_skip_target(fid):
                    matches.append((fid, 'param', t))
                    seen.add(fid)
        # DI expansion
        if t in di_map:
            for impl in di_map[t]:
                if impl in class_idx:
                    for fid in class_idx[impl]:
                        if fid not in seen and fid != rec['id'] and not should_skip_target(fid):
                            matches.append((fid, 'di', f"{t}->{impl}"))
                            seen.add(fid)

    # 2) new TypeName usages
    for typename in new_types:
        if typename in class_idx:
            for fid in class_idx[typename]:
                if fid not in seen and fid != rec['id'] and not should_skip_target(fid):
                    matches.append((fid, 'new', typename))
                    seen.add(fid)

    # 3) invocations: qualifier type or method name
    for expr, args in invs:
        parts = expr.split('.')
        method_or_type = parts[-1]
        qualifier = parts[0] if len(parts) > 1 else None
        if qualifier and qualifier[0].isupper() and qualifier in class_idx:
            for fid in class_idx[qualifier]:
                if fid not in seen and fid != rec['id'] and not should_skip_target(fid):
                    matches.append((fid, 'qualifier', qualifier))
                    seen.add(fid)
        if method_or_type in method_idx:
            for fid in method_idx[method_or_type]:
                if fid not in seen and fid != rec['id'] and not should_skip_target(fid):
                    matches.append((fid, 'method', method_or_type))
                    seen.add(fid)
        for arg in [a.strip() for a in args.split(',') if a.strip()]:
            if arg in var_map:
                t = var_map[arg]
                if t in class_idx:
                    for fid in class_idx[t]:
                        if fid not in seen and fid != rec['id'] and not should_skip_target(fid):
                            matches.append((fid, 'argvar', t))
                            seen.add(fid)

    # 4) using -> namespace matches
    # gather using candidates first; optionally include them only if also matched by other heuristics
    using_candidates = []
    for using in rec.get('usings', []):
        for fid in match_using_to_file_ids(using, ns_to_ids):
            if fid not in seen and fid != rec['id']:
                try:
                    target = next((r for r in globals().get('records', []) if r['id'] == fid), None)
                except Exception:
                    target = None
                skip = False
                if target:
                    fname = Path(target['path']).name.lower()
                    for kw in USING_IGNORE_KEYWORDS:
                        if kw in fname:
                            skip = True
                            break
                    if not skip:
                        for pat in EXCLUDE_FILENAME_PATTERNS:
                            if pat and pat in fname:
                                skip = True
                                break
                if not skip:
                    using_candidates.append((fid, using))

    # decide whether to add using candidates: if NO_USING_ONLY is True, only add those which
    # have already been matched by another heuristic (i.e., are in 'seen'). Otherwise add all.
    for fid, using in using_candidates:
        if fid in seen or not NO_USING_ONLY:
            matches.append((fid, 'using', using))
            seen.add(fid)

    # 5) filename fallback: if a referenced type name wasn't found, look for filename match
    # (cheap heuristic) - check param_field_types and new_types
    for t in set(list(param_field_types) + list(new_types)):
        if t not in class_idx:
            # attempt find a file whose filename matches TypeName.cs
            for r in globals().get('records', []):
                fname = Path(r['path']).name
                if fname.lower() == f"{t.lower()}.cs":
                    fid = r['id']
                    if fid not in seen and fid != rec['id'] and not should_skip_target(fid):
                        matches.append((fid, 'filename', t))
                        seen.add(fid)

    return rec['id'], rec['relpath'], matches


def find_variable_type_map(text: str):
    """Heuristic map of local variable name -> type by scanning 'var name = new Type' or 'Type name ='"""
    var_map = {}
    # var name = new Type(...)
    for m in re.finditer(r"\bvar\s+([A-Za-z0-9_]+)\s*=\s*new\s+([A-Za-z0-9_]+)", text):
        var_map[m.group(1)] = m.group(2)
    # Type name = new Type2(...)
    for m in re.finditer(r"\b([A-Za-z0-9_]+)\s+([A-Za-z0-9_]+)\s*=\s*new\s+([A-Za-z0-9_]+)", text):
        # prefer explicit typed var
        var_map[m.group(2)] = m.group(1)
    return var_map


def find_field_and_param_types(file_path: Path):
    """Return a set of type names referenced in field declarations and method parameter lists."""
    types = set()
    try:
        text = file_path.read_text(encoding='utf-8')
        # strip comments to avoid false matches coming from commented-out code
        text = strip_comments(text)
    except Exception:
        return types

    # fields: look for common field declaration patterns like 'private readonly IUsersService _usersService;'
    for m in re.finditer(r"\b(?:public|private|protected|internal|static|readonly|volatile|const)\s+([A-Za-z0-9_<>.,\s\[\]]+)\s+[A-Za-z0-9_]+\s*(?:=|;)", text):
        t = m.group(1).strip()
        # take first token (strip generics)
        tshort = re.sub(r"<.*>$", "", t).split()[-1]
        if tshort:
            types.add(tshort)

    # method parameter lists: capture parameter lists from method signatures
    for m in re.finditer(r"\b(?:public|private|protected|internal|static|async|protected internal|internal protected)\s+[A-Za-z0-9_<>,\s\[\]]+\s+[A-Za-z0-9_]+\s*\(([^)]*)\)", text):
        plist = m.group(1)
        if not plist:
            continue
        for p in plist.split(','):
            p = p.strip()
            if not p:
                continue
            # remove attributes like [FromBody]
            p = re.sub(r"\[[^\]]+\]", "", p).strip()
            # remove modifiers (ref/out/in/params)
            p = re.sub(r"\b(ref|out|in|params)\b", "", p).strip()
            parts = p.split()
            if len(parts) >= 2:
                t = parts[0]
                tshort = re.sub(r"<.*>$", "", t)
                types.add(tshort)
    return types


def extract_invocations(text: str):
    """Return list of invocation expressions (raw expression before '(') and argument list strings"""
    invocations = []
    for m in re.finditer(r"([A-Za-z0-9_\.]+)\s*\(([^)]*)\)", text):
        expr = m.group(1)
        args = m.group(2)
        invocations.append((expr, args))
    return invocations


def build_namespace_index(records):
    ns_to_ids = defaultdict(list)
    for rec in records:
        for ns in rec['declared_namespaces']:
            ns_to_ids[ns].append(rec['id'])
    return ns_to_ids


def build_di_registration_map(src_root: Path, ignore_globs, ignore_regexes):
    """Scan source files for DI registrations like AddScoped<IService, Service>() or AddScoped(typeof(IService), typeof(Service))
    Returns mapping interfaceShortName -> list of implementation short names.
    """
    di_map = defaultdict(list)
    pattern_generic = re.compile(r"\bAdd(?:Scoped|Transient|Singleton)\s*<\s*([A-Za-z0-9_\.<>]+)\s*,\s*([A-Za-z0-9_\.<>]+)\s*>", re.IGNORECASE)
    pattern_typeof = re.compile(r"\bAdd(?:Scoped|Transient|Singleton)\s*\(\s*typeof\(\s*([A-Za-z0-9_\.<>]+)\s*\)\s*,\s*typeof\(\s*([A-Za-z0-9_\.<>]+)\s*\)\s*\)", re.IGNORECASE)

    def is_ignored(p: Path):
        try:
            rel = str(p.relative_to(src_root)).replace('\\', '/')
        except Exception:
            rel = str(p).replace('\\', '/')
        for g in ignore_globs:
            if fnmatch.fnmatch(rel, g) or fnmatch.fnmatch(str(p).replace('\\', '/'), g):
                return True
        for rg in ignore_regexes:
            if rg.search(rel) or rg.search(str(p)):
                return True
        return False

    for p in src_root.rglob('*.cs'):
        if p.is_file() and not is_ignored(p):
            try:
                text = p.read_text(encoding='utf-8')
            except Exception:
                continue
            for m in pattern_generic.finditer(text):
                iface = m.group(1).split('.')[-1]
                impl = m.group(2).split('.')[-1]
                if impl not in di_map[iface]:
                    di_map[iface].append(impl)
            for m in pattern_typeof.finditer(text):
                iface = m.group(1).split('.')[-1]
                impl = m.group(2).split('.')[-1]
                if impl not in di_map[iface]:
                    di_map[iface].append(impl)

    return di_map


def match_using_to_file_ids(using, ns_to_ids):
    """Return file ids whose declared namespace matches the using.

    In normal mode we allow parent->child namespace matches (e.g. using Foo; will match Foo.Bar).
    If GLOBAL_STRICT_USINGS is True we use a conservative rule: only declared namespace == using or
    declared namespace startswith(using + '.') (i.e. using is a parent or equal). We explicitly
    remove the reverse check which was causing some noisy matches.
    """
    matches = set()
    for decl_ns, ids in ns_to_ids.items():
        if GLOBAL_STRICT_USINGS:
            if decl_ns == using or decl_ns.startswith(using + '.'):
                matches.update(ids)
        else:
            # legacy permissive behavior (keeps prior matching semantics)
            if decl_ns == using or decl_ns.startswith(using + '.') or using.startswith(decl_ns + '.'):
                matches.update(ids)
    return matches


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--source-root', help='Source root to scan (will prompt if omitted)')
    parser.add_argument('--extensions', default='.cs', help='Comma-separated file extensions to include for namespace/usings scanning (default: .cs)')
    parser.add_argument('--output', default='asts_enhanced/file_imports_from_source.xlsx', help='Output XLSX file')
    parser.add_argument('--ignore-glob', action='append', default=[], help='Glob pattern to ignore (can be passed multiple times). Example: **/obj/**')
    parser.add_argument('--ignore-regex', action='append', default=[], help='Regex pattern to ignore (can be passed multiple times).')
    parser.add_argument('--autosize', action='store_true', help='Enable autosize columns (disabled by default for speed)')
    parser.add_argument('--workers', type=int, default=max(1, multiprocessing.cpu_count() - 1), help='Number of worker processes to use for parsing')
    parser.add_argument('--strict-usings', action='store_true', help='Only match using->namespace conservatively (reduces noisy using matches)')
    parser.add_argument('--exclude-filename-pattern', action='append', default=[], help='Substring pattern to exclude matching filenames (case-insensitive). Can be passed multiple times.')
    parser.add_argument('--no-using-only', action='store_true', help="Don't include imports that are only matched via 'using' (keeps only imports with other match reasons)")
    args = parser.parse_args()

    src_root = args.source_root
    if not src_root:
        # prompt interactively and offer current working directory as default
        try:
            default_root = os.getcwd()
            resp = input(f'Enter source root path [{default_root}]: ').strip()
            src_root = resp or default_root
        except Exception:
            print('No source root provided and cannot prompt. Use --source-root.')
            return 2
    src_root = Path(src_root)
    if not src_root.exists():
        print('Source root does not exist:', src_root)
        return 2

    # apply CLI-driven globals
    global GLOBAL_STRICT_USINGS
    GLOBAL_STRICT_USINGS = bool(args.strict_usings)
    global EXCLUDE_FILENAME_PATTERNS
    EXCLUDE_FILENAME_PATTERNS = [p.lower() for p in args.exclude_filename_pattern]
    global NO_USING_ONLY
    NO_USING_ONLY = bool(args.no_using_only)

    exts = [e.strip().lower() for e in args.extensions.split(',') if e.strip()]
    output = Path(args.output)

    # collect files
    default_ignore = ['**/obj/**', '**/bin/**']
    ignore_globs = list(default_ignore) + args.ignore_glob
    ignore_regexes = [re.compile(r) for r in args.ignore_regex]

    def is_ignored_path(p: Path):
        # ignore by glob against relative path, and by regex against relative or absolute path
        try:
            rel = str(p.relative_to(src_root)).replace('\\', '/')
        except Exception:
            rel = str(p).replace('\\', '/')
        # glob patterns
        for pattern in ignore_globs:
            if fnmatch.fnmatch(rel, pattern) or fnmatch.fnmatch(str(p).replace('\\', '/'), pattern):
                return True
        # regex patterns
        for rg in ignore_regexes:
            if rg.search(rel) or rg.search(str(p)):
                return True
        return False

    all_files = [p for p in src_root.rglob('*') if p.is_file() and not is_ignored_path(p)]
    # sheet1: file types
    ext_counter = Counter()
    for p in all_files:
        ext = p.suffix.lower() if p.suffix else '(no_ext)'
        ext_counter[ext] += 1

    # for files of interest (.cs etc), extract namespaces/usings
    source_files = [p for p in all_files if p.suffix.lower() in exts]
    records = []
    for idx, p in enumerate(sorted(source_files), start=1):
        rel = str(p.relative_to(src_root))
        nss, us = find_namespaces_and_usings(p)
        records.append({
            'id': idx,
            'path': str(p),
            'relpath': rel,
            'declared_namespaces': nss,
            'usings': us
        })

    # Precompute and cache common regex-derived info per file to avoid re-reading/parsing during import detection.
    # This speeds up the sequential import loop significantly.
    for rec in records:
        try:
            text_raw = Path(rec['path']).read_text(encoding='utf-8')
            text = strip_comments(text_raw)
        except Exception:
            text = ''
        rec['text'] = text
        rec['var_map'] = find_variable_type_map(text)
        try:
            rec['param_field_types'] = find_field_and_param_types(Path(rec['path']))
        except Exception:
            rec['param_field_types'] = set()
        # new TypeName usages
        rec['new_types'] = set(m.group(1) for m in re.finditer(r"new\s+([A-Za-z0-9_]+)", text))
        rec['invocations'] = extract_invocations(text)

    # build declaration indexes for classes and methods (parallelized)
    class_idx, method_idx = build_decl_indexes_parallel(records, workers=args.workers)

    ns_to_ids = build_namespace_index(records)
    # build DI registration map (interface -> implementations)
    di_map = build_di_registration_map(src_root, ignore_globs, ignore_regexes)

    # Build namespace index (used for explicit using->file matches)
    ns_to_ids = build_namespace_index(records)

    # create workbook. Use write-only mode by default for speed; enable regular workbook only if autosize requested
    if args.autosize:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = sanitize_sheet_name('FileTypes')
    else:
        # create a write-only workbook for much lower memory use and faster writes
        wb = Workbook(write_only=True)
        ws1 = wb.create_sheet(title=sanitize_sheet_name('FileTypes'))
    ws1.append(['Extension', 'Count'])
    for k, v in sorted(ext_counter.items(), key=lambda x: (-x[1], x[0])):
        ws1.append([k, v])
    if args.autosize:
        try:
            autosize(ws1, cols=[1,2])
        except Exception:
            pass

    ws2 = wb.create_sheet(title=sanitize_sheet_name('Files'))
    # Do not include absolute path column as requested
    ws2.append(['FileID', 'RelPath', 'DeclaredNamespaces', 'Usings'])
    for r in records:
        ws2.append([r['id'], r['relpath'], '; '.join(r['declared_namespaces']), '; '.join(r['usings'])])
    if args.autosize:
        try:
            autosize(ws2)
        except Exception:
            pass

    ws3 = wb.create_sheet(title=sanitize_sheet_name('Imports'))
    # Add columns to show WHY a file was included (diagnostic)
    ws3.append(['FileID', 'RelPath', 'ImportedFileID', 'ImportedRelPath', 'MatchedBy', 'MatchedSymbol'])

    # Heuristic: for each file, find referenced files by the same heuristics as before.
    # Process sequentially to preserve deterministic results (parallel workers were causing incorrect/misaligned outputs).
    total = len(records)
    for idx, rec in enumerate(records, start=1):
        if idx % 50 == 0 or idx == total:
            print(f'Processing imports: {idx}/{total}')
        try:
            fid, rel, imported = process_record_imports((rec, class_idx, method_idx, di_map, ns_to_ids))
        except Exception:
            continue
        if not imported:
            ws3.append([fid, rel, '', '', '', ''])
        else:
            for iid, matched_by, matched_sym in imported:
                imp = records[iid - 1]
                ws3.append([fid, rel, iid, imp['relpath'], matched_by, matched_sym])
    if args.autosize:
        try:
            autosize(ws3)
        except Exception:
            pass

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    print('Wrote', output)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
